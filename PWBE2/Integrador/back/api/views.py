from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.parsers import MultiPartParser
from rest_framework.generics import ListCreateAPIView, RetrieveUpdateDestroyAPIView, CreateAPIView
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import load_workbook, Workbook
from .serializers import *
from .models import Sensores, Ambientes, Historico
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
from .filters import *
from django.http import HttpResponse
import datetime

# --- SENSORES ---

class SensoresView(ListCreateAPIView):
    queryset = Sensores.objects.all()
    serializer_class = SensoresSerializer 
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_class = SensoresFilter

class SensoresDetailView(RetrieveUpdateDestroyAPIView):
    queryset = Sensores.objects.all()
    serializer_class = SensoresSerializer 
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_class = SensoresFilter

# --- AMBIENTES ---

class AmbientesView(ListCreateAPIView):
    queryset = Ambientes.objects.all()
    serializer_class = AmbientesSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_class = AmbienteFilter

class AmbientesDetailView(RetrieveUpdateDestroyAPIView):
    queryset = Ambientes.objects.all()
    serializer_class = AmbientesSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_class = AmbienteFilter

# --- HISTÓRICO ---

class HistoricoView(ListCreateAPIView):
    queryset = Historico.objects.all()
    serializer_class = HistoricoSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_class = HistoricoFilter

class HistoricoDetailView(RetrieveUpdateDestroyAPIView):
    queryset = Historico.objects.all()
    serializer_class = HistoricoSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_class = HistoricoFilter

# --- USUÁRIOS ---

class SignUpView(CreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [AllowAny]

class LoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            # login(session) só funciona se tiver SessionMiddleware no Django, se for API puro, melhor usar JWT ou Token
            login(request, user)
            return Response({'message': 'Login realizado com sucesso!'})
        else:
            return Response({'error': 'Credenciais inválidas'}, status=400)


# --- IMPORTAÇÃO DE PLANILHA XLSX (SENSORES) ---

class UploadXLSXSensoresView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'Arquivo não enviado'}, status=400)

        wb = load_workbook(filename=file_obj)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            sensor, mac_address, unidade_med, latitude, longitude, status = row
            try:
                Sensores.objects.create(
                    sensor=sensor.strip(),
                    mac_address=mac_address.strip(),
                    unidade_med=unidade_med.strip(),
                    latitude=float(latitude),
                    longitude=float(longitude),
                    status=status.strip().lower() in ['ativo', 'true', '1']
                )
            except Exception as e:
                print(f"Erro ao salvar sensor: {e}")
                continue

        return Response({'message': 'Sensores importados com sucesso!'})
    

class UploadXLSXAmbientesView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'Arquivo não enviado'}, status=400)

        wb = load_workbook(filename=file_obj)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            sig, descricao, ni, responsavel = row
            try:
                Ambientes.objects.create(
                    sig=sig,
                    descricao=descricao,
                    ni=ni,
                    responsavel=responsavel
                )
            except Exception as e:
                print(f"Erro ao salvar ambiente: {e}")
                continue

        return Response({'message': 'Ambientes importados com sucesso!'})

class UploadXLSXHistoricoView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'Arquivo não enviado'}, status=400)

        try:
            wb = load_workbook(filename=file_obj)
            ws = wb.active
        except Exception as e:
            return Response({'error': f'Erro ao ler o arquivo: {str(e)}'}, status=400)

        salvos = 0
        erros = []

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or len(row) != 4:
                erros.append(f"Linha {idx}: formato inválido: {row}")
                continue

            valor, timestamp_str, ambiente_id, sensor_id = row

            try:
                # Conversão de timestamp
                if isinstance(timestamp_str, str):
                    timestamp = datetime.strptime(timestamp_str, "%d/%m/%Y %H:%M")
                else:
                    timestamp = timestamp_str  # se o Excel já vier com datetime nativo

                # Busca pelos objetos relacionados
                sensor_obj = Sensores.objects.filter(id=int(sensor_id)).first()
                ambiente_obj = Ambientes.objects.filter(id=int(ambiente_id)).first()

                if not sensor_obj:
                    erros.append(f"Linha {idx}: sensor ID '{sensor_id}' não encontrado")
                    continue

                if not ambiente_obj:
                    erros.append(f"Linha {idx}: ambiente ID '{ambiente_id}' não encontrado")
                    continue

                # Criação do histórico
                Historico.objects.create(
                    valor=float(valor),
                    timestamp=timestamp,
                    sensor=sensor_obj,
                    ambiente=ambiente_obj
                )
                salvos += 1

            except Exception as e:
                erros.append(f"Linha {idx}: erro ao processar - {str(e)}")
                continue

        return Response({
            'message': f'{salvos} registros importados com sucesso!',
            'falhas': erros
        })


            

class ExportarXLSXSensoresView(APIView):
    def get(self, request):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="historico.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "Sensores"

        ws.append(['Valor', 'Timestamp', 'Sensor', 'Ambiente'])

        historico = Historico.objects.all()
        for item in historico:
            ws.append([
                item.valor,
                item.timestamp,
                item.sensor.sensor,
                item.ambiente.sig
            ])

        wb.save(response)
        return response
    

class ExportarXLSXAmbientesView(APIView):
    def get(self, request):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="ambientes.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "Sensores"

        ws.append(['Sigla', 'Descrição', 'NI', 'Responsável'])

        ambientes = Ambientes.objects.all()
        for ambiente in ambientes:
            ws.append([
                ambiente.sig,
                ambiente.descricao,
                ambiente.ni,
                ambiente.responsavel
            ])

        wb.save(response)
        return response
    

class ExportarXLSXHistoricoView(APIView):
    def get(self, request):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="historico.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "Histórico"

        ws.append(['Valor', 'Timestamp', 'Sensor', 'Ambiente'])

        historico = Historico.objects.all()
        for item in historico:
            ws.append([
                item.valor,
                item.timestamp,
                item.sensor.sensor,
                item.ambiente.sig
            ])

        wb.save(response)
        return response    


class SensoresFilterView(APIView):
    def get(self, request):
        queryset = Sensores.objects.all()
        serializer = SensoresSerializer(queryset, many=True)
        return Response(serializer.data)
    

class AmbientesFilterView(APIView):
    def get(self, request):
        queryset = Ambientes.objects.all()
        serializer = AmbientesSerializer(queryset, many=True)
        return Response(serializer.data)
    
class HistoricoFilterView(APIView):
    def get(self, request):
        queryset = Historico.objects.all()
        serializer = HistoricoSerializer(queryset, many=True)
        return Response(serializer.data)